import csv
import time
import numpy as np
from openpyxl import Workbook
from ListaODsimples import ListaLigada

tempo_inicial = time.time()

listaODpessoas = ListaLigada()

with open('OD_2017.csv') as csvfile:
    readCSV = csv.reader(csvfile, delimiter=',')

    i = 0

    for row in readCSV:
        if row[125].isnumeric():  # não colocar a primeira linha (info das tabelas)
            print("Adicionando à lista dado No:", i, "de coordenadas:", row[88], row[89], "e IDpessoa:", row[43])
            # Se local já existe na lista, apenas adicionará o IDPessoa no vetor "frequentadores" do nó
            listaODpessoas.append(int(row[88]), int(row[89]), int(row[43]))
            i += 1
            # Para testes
            #if i == 100000:
            #    break

max = 0

# Tratar casos 0,0
dado = listaODpessoas.getItem(0, 0)
splitDado = dado.split(",")
indexInvalido = int(splitDado[3])


# define maior valor possível de frequentadores em um local
for r in range(len(listaODpessoas)):
    print("Analisando frequência do dado No {0} da listaODpessoas".format(r))
    freq = listaODpessoas.contFreqIndex(r)
    if freq > max and r != indexInvalido:
        max = freq

# cria vetor contendo valores 0 em que as posições representam os frequentadores e o valor
# a quantidade de locais com este número de frequentadores
qtdLocais = np.zeros(max+1)
index = 0

# incrementa em 1 toda vez que a quantidade de frequentadores de um local for igual a da posição
# no vetor
for r in range(len(listaODpessoas)):
    print("Incrementando local com o dado No {0} da listaODpessoas".format(r))
    index = listaODpessoas.contFreqIndex(r)
    for i in range(max + 1):
        if i == index:
            qtdLocais[i] += 1

arquivo_excel = Workbook()
planilha1 = arquivo_excel.active
planilha1.title = "Histograma Origem-Destino"

planilha1.cell(row=1, column=1, value="QUANTIDADE DE FREQUENTADORES")
planilha1.cell(row=1, column=2, value="QUANT. DE LOCAIS COM ESTE NÚMERO DE FREQUENTADORES")

for r in range(2, max + 3):  # era 1
    planilha1.cell(row=r, column=1, value=r - 2)  # era 1
    planilha1.cell(row=r, column=2, value=qtdLocais[r - 2])

arquivo_excel.save("Histograma Origem-Destino.xlsx")

print("")
print("O número de frequentadores máximo encontrado foi:", max)
print("Dados armazenados no excel!")

print("")
print("--- %s segundos ---" % (time.time() - tempo_inicial))




