from Grafo import Grafo
import numpy as np
from openpyxl import Workbook

# Tamanho do arranjo original: 86318
print("Criando grafo...")
grafoODpessoas = Grafo(86319)

for i in range(86319):
    grafoODpessoas.novo_Vertice(i)

# Descarta as primeiras duas linhas do arquivo
with open('OD_graph.txt') as f:
    f.readline()
    f.readline()
    restante = f.readlines()

i = 0
for line in restante:  # comeca na linha 3
    print("Adicionando dados da linha:", i)
    dadoSplit = line.split(" ")
    grafoODpessoas.novo_NoAdj(int(dadoSplit[0]), int(dadoSplit[1]))
    i += 1

max = 0
# define maior valor possível de grau dos vértices
for r in range(86319):
    print("Analisando grau do dado No {0} da grafoODpessoas".format(r))
    grau = grafoODpessoas.contar_GrauVertice(r)
    if grau > max:
        max = grau


# cria vetor contendo valores 0 em que as posições representam os graus e o valor
# a quantidade de vértices com este número de graus
qtdVertporGrau = np.zeros(max+1)
index = 0

# incrementa em 1 toda vez que o grau de um vértice for igual a da posição no vetor
for r in range(86319):
    print("Incrementando grau com o dado No {0} do grafoODpessoas".format(r))
    index = grafoODpessoas.contar_GrauVertice(r)
    for i in range(max + 1):
        if i == index:
            qtdVertporGrau[i] += 1


arquivo_excel = Workbook()
planilha1 = arquivo_excel.active
planilha1.title = "Histograma Grau por Vértice"

planilha1.cell(row=1, column=1, value="GRAU")
planilha1.cell(row=1, column=2, value="QUANT. DE VERTICES COM ESTE GRAU")

for r in range(2, max + 3):  # era 1
    planilha1.cell(row=r, column=1, value=r - 2)  # era 1
    planilha1.cell(row=r, column=2, value=qtdVertporGrau[r - 2])

arquivo_excel.save("Histograma Grau por Vértice.xlsx")

print("")
print("A quantidade de arestas total é:", grafoODpessoas.contar_arestasGrafo())
print("Maior grau encontrado:", max)
print("Dados armazenados no excel!")



