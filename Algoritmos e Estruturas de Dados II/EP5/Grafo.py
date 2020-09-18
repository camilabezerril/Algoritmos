import random
from random import uniform
from collections import deque
from openpyxl import Workbook


class Vertice:
    def __init__(self, IDpessoa):
        self.id = IDpessoa
        self.vizinhos = list()

    def acrescentar_Vizinho(self, v):
        if v not in self.vizinhos:
            self.vizinhos.append(v)
            self.vizinhos.sort()


class Grafo:
    todosVertices = {}
    verticesValidos = []
    marked = []
    SIRcomp = []
    count = 0
    probContagio = 0
    probRecuperar = 0

    def adicionar_Vertice(self, v):
        if isinstance(v, Vertice) and v.id not in self.todosVertices:
            print("Adicionando vértice {0} à lista".format(v.id))
            self.todosVertices[v.id] = v
            return True
        else:
            return False

    def adicionar_NoAdj(self, u, v):
        if u in self.todosVertices and v in self.todosVertices:
            for key, value in self.todosVertices.items():
                if key == u:
                    value.acrescentar_Vizinho(v)
                    if u not in self.verticesValidos:
                        self.verticesValidos.append(u)
                if key == v:
                    value.acrescentar_Vizinho(u)
                    if v not in self.verticesValidos:
                        self.verticesValidos.append(v)
            return True
        else:
            return False

    def imprimir_Grafo(self):
        for key in sorted(list(self.todosVertices.keys())):
            print(str(key) + str(self.todosVertices[key].vizinhos))

    def mostrar_Vertice(self, key):
        print(str(key) + str(self.todosVertices[key].vizinhos))

    def contar_Grau(self, key):
        # Não contar o vértice inicial, dando assim o número de arestas
        return len(self.todosVertices[key].vizinhos)

    def contar_ArestasGrafo(self):
        total = 0
        for key in sorted(list(self.todosVertices.keys())):
            total = total + len(self.todosVertices[key].vizinhos)
        return total

    def create_SIR(self, max):
        for x in range(max + 1):
            self.SIRcomp.append('S')

    def simulacaoSIR(self, caso):
        pacienteZero = random.choice(self.verticesValidos)

        if caso == 1:  # Caso em que quase ninguém se precaveu
            self.probContagio = 0.9
            self.probRecuperar = 0.1
        elif caso == 2:  # Caso em que alguns estão precavidos
            self.probContagio = 0.8
            self.probRecuperar = 0.25
        elif caso == 3:  # Caso em que todos estão precavidos no local
            self.probContagio = 0.6
            self.probRecuperar = 0.4
        else:
            print("Caso inválido")

        self.create_SIR(len(self.todosVertices))
        self.SIRcomp[pacienteZero] = 'I'

        stack = []
        stack.append(pacienteZero)

        infectados = 1
        recuperados = 0
        suscetiveis = len(self.verticesValidos) - 1

        arquivo_excel = Workbook()
        planilha1 = arquivo_excel.active
        planilha1.title = "Histograma modeloSIR caso"+str(caso)

        planilha1.cell(row=1, column=1, value="PASSO")
        planilha1.cell(row=1, column=2, value="SUSCETÍVEL")
        planilha1.cell(row=1, column=3, value="INFECTADO")
        planilha1.cell(row=1, column=4, value="RECUPERADO")

        count = 1
        while infectados != 0:
            auxRec = []
            auxInfec = []
            count += 1
            for v in stack:
                if self.acasoR_SIR(v):
                    auxRec.append(v)
                    infectados -= 1
                    recuperados += 1
                else:
                    for w in self.todosVertices[v].vizinhos:
                        if self.acasoI_SIR(w):
                            auxInfec.append(w)
                            infectados += 1
                            suscetiveis -= 1

            for i in range(len(auxRec)):
                auxInd = -1
                for j in range(len(stack)):
                    if stack[j] == auxRec[i]:
                        auxInd = j
                stack.pop(auxInd)

            for i in range(len(auxInfec)):
                stack.append(auxInfec[i])

            planilha1.cell(row=count, column=1, value=count-1)
            planilha1.cell(row=count, column=2, value=suscetiveis)
            planilha1.cell(row=count, column=3, value=infectados)
            planilha1.cell(row=count, column=4, value=recuperados)


        arquivo_excel.save("Histograma modeloSIR caso"+str(caso)+".xlsx")
        print("Dados armazenados no excel!")

        print("")
        print("O paciente zero foi:", pacienteZero)
        print("Número de infectados:", infectados)
        print("Número de recuperados:", recuperados)
        print("Número de suscetíveis:", suscetiveis)

    def acasoI_SIR(self, v):
        chanceI = round(uniform(0, 1), 2)  # uniform: faixa de ponto flutuante

        print("")
        print("Probabilidade de se infectar:", self.probContagio)
        print("Probabilidade tirada pela pessoa {0}: {1}".format(v, chanceI))

        if self.SIRcomp[v] == 'S' and chanceI <= self.probContagio:
            self.SIRcomp[v] = 'I'
            print("Esta pessoa se infectou.")
            return True
        else:
            print("Esta pessoa não se infectou.")
            return False

    def acasoR_SIR(self, v):
        chanceR = round(uniform(0, 1), 2)  # uniform: faixa de ponto flutuante

        print("")
        print("Probabilidade de se recuperar:", self.probRecuperar)
        print("Probabilidade tirada pela pessoa {0}: {1}".format(v, chanceR))

        if chanceR <= self.probRecuperar:
            self.SIRcomp[v] = 'R'
            print("Esta pessoa se recuperou.")
            return True
        else:
            print("Esta pessoa não se recuperou.")
            return False